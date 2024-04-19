import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
import os

def input_excel(new_form,f_path,tmp_path):
    """データフレームをエクセルファイル（室長パトロール結果_雛型）に入力
       生成したexcelは「環境・バイオ研究室.xlsx」として同じ階層に保存
    Parameter:
    ---------
    new_form: dataframe
    machine_name: str
    f_path: str
        現在位置のパス
    temp_path: str
        雛型エクセルデータのパス
    """

    list_d = [list(new_form.loc[i,:]) for i in range(len(new_form))]

    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    
    wb = load_workbook(tmp_path)
    ws = wb['Sheet1'] 

    # 空白列の確認
    blank_cell_row_ = ""

    for row_num in range(17,100):
        cell_value = ws.cell(row=row_num,column=1).value
        if cell_value == None:
            blank_cell_row_ = row_num
            break
    blank_cell_row = int(blank_cell_row_)

    for y, row in enumerate(list_d):
        for x, cells in enumerate(row):
            ws.cell(row=blank_cell_row+y,column=1+x,value=list_d[y][x])
            if ws.cell(row=blank_cell_row+y,column=1+x).value:
                ws.cell(row=blank_cell_row+y,column=1+x).border= border           

    wb.save(f'{f_path}/環境・バイオ研究室.xlsx')


f_path = os.getcwd()
tmp_path = f'{f_path}/雛型/雛型.xlsx'

df = pd.read_csv('室長パトロール結果.csv')
df2 = df.loc[:,["field_0","Title","OData__x90e8__x5c4b__x540d_","field_3","field_4","field_2"]]
df2.columns = [["点検月","指摘事項","対象実験室","担当者","対応結果","対応詳細",]]

input_excel(df2,f_path,tmp_path)

os.remove("室長パトロール.csv")